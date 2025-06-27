import openpyxl
import re
from openpyxl import load_workbook

wb1 = openpyxl.load_workbook(r'C:\Users\rchrd\OneDrive\Desktop\openpyxl\cleaned.xlsx.xlsx')
wb2 = openpyxl.load_workbook(r'C:\Users\rchrd\OneDrive\Desktop\openpyxl\hops.xlsx.xlsx')

ws1 = wb1.active
ws2 = wb2.active

last_row = ws1.max_row

### COPY ROWS FROM CLEANED TO HOPS WORKBOOKS
for row in range(2, last_row + 1):
    ws2[f'C{row}'].value = ws1[f'C{row}'].value
for row in range(2, last_row + 1):
    ws2[f'N{row}'].value = ws1[f'F{row}'].value
### COPY THE 1ST 10 CHARACTERS TO NEW ROWS...
row = 2

# Loop through rows until there's no more data in column C or N
while True:
    c_val = ws2[f'C{row}'].value
    n_val = ws2[f'N{row}'].value

    # If both are empty, stop the loop
    if c_val is None and n_val is None:
        break

    # Copy first 10 characters from column C to G
    if c_val is not None:
        ws2[f'G{row}'] = str(c_val)[:10]

    # Copy first 10 characters from column N to K
    if n_val is not None:
        ws2[f'K{row}'] = str(n_val)[:10]

    row += 1

### APPEND THE POD PATCH LOCATIONS:
# Start from row 2
row = 2

while True:
    k_val = ws2[f'K{row}'].value

    # Stop if no value in column K
    if k_val is None:
        break

    k_val_str = str(k_val)
    g_val = ws2[f'G{row}'].value or ''

    # Match values that start with 2602 or 2606 using regex
    if re.match(r'^RM2602', k_val_str):
        ws2[f'G{row}'] = str(g_val) + 'C10.U45'
    elif re.match(r'^RM2606', k_val_str):
        ws2[f'G{row}'] = str(g_val) + 'C11.U45'

    row += 1

### APPENDING NETWORK PATCH CAB AND U HEIGHTS:
# Mapping of G prefixes to K suffixes
g_prefix_to_k_append = {
    'RM2302-R1': 'C17.U42',
    'RM2302-R2': 'C17.U37',
    'RM2302-R3': 'C17.U32',
    'RM2302-R4': 'C17.U27',
    'RM2302-R5': 'C17.U22',
    'RM2302-R6': 'C17.U17',

    'RM2802-R1': 'C19.U42',
    'RM2802-R2': 'C19.U37',
    'RM2802-R3': 'C19.U32',
    'RM2802-R4': 'C19.U27',
    'RM2802-R5': 'C19.U22',
    'RM2802-R6': 'C19.U17',

    'RM2304-R1': 'C21.U42',
    'RM2304-R2': 'C21.U37',
    'RM2304-R3': 'C21.U32',
    'RM2304-R4': 'C21.U27',
    'RM2304-R5': 'C21.U22',
    'RM2304-R6': 'C21.U17',

    'RM2602-R1': 'C26.U32',

    'RM2606-R1': 'C26.U32'
}
# Start from row 2
row = 2

while True:
    g_val = ws2[f'G{row}'].value

    if g_val is None:
        break

    g_val_str = str(g_val)
    k_val = ws2[f'K{row}'].value or ''

    # Check each prefix and append to column K accordingly
    for prefix, append_value in g_prefix_to_k_append.items():
        if g_val_str.startswith(prefix):
            ws2[f'K{row}'] = str(k_val) + append_value
            break  # Only one match per row

    row += 1

wb2.save('COMPLETED.xlsx')

